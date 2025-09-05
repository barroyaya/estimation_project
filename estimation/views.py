# estimation/views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.db.models import Q, Sum
from django.core.paginator import Paginator
from .models import *
import json


def index(request):
    """Page d'accueil pour les clients"""
    projets_actifs = Projet.objects.filter(actif=True).count()
    categories = Categorie.objects.all().count()
    elements = Element.objects.filter(actif=True).count()

    context = {
        'stats': {
            'projets': projets_actifs,
            'categories': categories,
            'elements': elements,
        }
    }
    return render(request, 'client/index.html', context)


def project_selection(request):
    """Sélection d'un projet existant ou création d'un nouveau"""
    projets = Projet.objects.filter(actif=True)

    if request.method == 'POST':
        if 'nouveau_projet' in request.POST:
            nom_projet = request.POST.get('nom_projet')
            client = request.POST.get('client', '')
            if nom_projet:
                projet = Projet.objects.create(nom=nom_projet, client=client)
                EstimationSummary.objects.create(projet=projet)
                request.session['projet_id'] = projet.id
                messages.success(request, f'Projet "{nom_projet}" créé avec succès!')
                return redirect('category_selection')
        elif 'projet_existant' in request.POST:
            projet_id = request.POST.get('projet_id')
            if projet_id:
                request.session['projet_id'] = int(projet_id)
                return redirect('category_selection')

    return render(request, 'client/project_selection.html', {'projets': projets})


def category_selection(request):
    """Sélection de catégorie pour un projet"""
    projet_id = request.session.get('projet_id')
    if not projet_id:
        return redirect('project_selection')

    projet = get_object_or_404(Projet, id=projet_id)
    categories = Categorie.objects.all()

    # Grouper par type
    categories_groupees = {}
    for cat in categories:
        type_cat = cat.get_type_categorie_display()
        if type_cat not in categories_groupees:
            categories_groupees[type_cat] = []
        categories_groupees[type_cat].append(cat)

    context = {
        'projet': projet,
        'categories_groupees': categories_groupees
    }

    if request.method == 'POST':
        categorie_id = request.POST.get('categorie_id')
        if categorie_id:
            return redirect('item_selection', categorie_id=categorie_id)

    return render(request, 'client/category_selection.html', context)


# imports en haut du fichier views.py (ajoute Unite)
from .models import Projet, Categorie, Discipline, Element, EstimationElement, EstimationSummary, DemandeElement, Unite
from django.db.models import Q
from django.core.paginator import Paginator
from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect, render

def item_selection(request, categorie_id):
    """Sélection d'éléments dans une catégorie"""
    projet_id = request.session.get('projet_id')
    if not projet_id:
        return redirect('project_selection')

    projet = get_object_or_404(Projet, id=projet_id)
    categorie = get_object_or_404(Categorie, id=categorie_id)

    # Filtres
    search_query = request.GET.get('search', '')
    discipline_id = request.GET.get('discipline', '')

    # ⚠️ Précharger discipline + unite (FK)
    elements = (Element.objects
                .filter(categorie=categorie, actif=True)
                .select_related('discipline', 'unite'))

    if search_query:
        elements = elements.filter(
            Q(designation__icontains=search_query) |
            Q(caracteristiques__icontains=search_query)
        )

    if discipline_id:
        elements = elements.filter(discipline_id=discipline_id)

    # Pagination
    paginator = Paginator(elements, 20)
    page_number = request.GET.get('page')
    page_elements = paginator.get_page(page_number)

    # Éléments déjà sélectionnés
    elements_selectionnes = EstimationElement.objects.filter(
        projet=projet,
        element__categorie=categorie
    ).values_list('element_id', flat=True)

    # Récupérer les demandes personnalisées (⚠️ précharger unite aussi)
    demandes_personnalisees = (DemandeElement.objects
        .filter(projet=projet, categorie=categorie)
        .select_related('discipline', 'unite')
        .order_by('-date_demande'))

    disciplines = Discipline.objects.filter(
        id__in=elements.values_list('discipline_id', flat=True).distinct()
    )

    context = {
        'projet': projet,
        'categorie': categorie,
        'elements': page_elements,
        'disciplines': disciplines,
        'elements_selectionnes': list(elements_selectionnes),
        'demandes_personnalisees': demandes_personnalisees,
        'search_query': search_query,
        'discipline_selectionnee': discipline_id,
    }

    if request.method == 'POST':
        # Gestion du formulaire d'élément personnalisé
        if 'nouveau_element' in request.POST:
            designation = request.POST.get('designation_personnalisee', '').strip()
            discipline_id = request.POST.get('discipline_personnalisee')
            caracteristiques = request.POST.get('caracteristiques_personnalisees', '').strip()
            unite_code = request.POST.get('unite_personnalisee', 'u')  # ex: 'ml', 'm2', ...
            quantite = request.POST.get('quantite_personnalisee', 1)

            if designation and discipline_id:
                try:
                    discipline = get_object_or_404(Discipline, id=discipline_id)
                    quantite_float = float(quantite) if quantite else 1.0
                    unite_obj = Unite.objects.filter(code=unite_code).first() or Unite.objects.get(code='u')

                    DemandeElement.objects.create(
                        projet=projet,
                        categorie=categorie,
                        discipline=discipline,
                        designation=designation,
                        caracteristiques=caracteristiques,
                        unite=unite_obj,  # ✅ FK, pas un code texte
                        quantite=quantite_float
                    )

                    messages.success(request, f'Votre demande pour "{designation}" a été envoyée avec succès!')
                    return redirect('item_selection', categorie_id=categorie_id)

                except (ValueError, TypeError):
                    messages.error(request, 'Veuillez entrer une quantité valide.')
                except Exception:
                    messages.error(request, 'Une erreur est survenue lors de l\'envoi de votre demande.')
            else:
                messages.error(request, 'Veuillez remplir tous les champs obligatoires (désignation et discipline).')

        # Gestion du formulaire d'éléments standards
        else:
            element_ids = request.POST.getlist('elements')
            quantites = request.POST.getlist('quantites')

            if element_ids:
                try:
                    EstimationElement.objects.filter(
                        projet=projet,
                        element__categorie=categorie
                    ).delete()

                    elements_ajoutes = 0
                    for i, element_id in enumerate(element_ids):
                        if element_id:
                            element = Element.objects.get(id=element_id)
                            quantite = float(quantites[i]) if quantites[i] else 1.0

                            EstimationElement.objects.create(
                                projet=projet,
                                element=element,
                                quantite=quantite
                            )
                            elements_ajoutes += 1

                    summary, _ = EstimationSummary.objects.get_or_create(projet=projet)
                    summary.calculer_totaux()

                    messages.success(request, f'{elements_ajoutes} élément(s) ajouté(s) à votre estimation!')
                    return redirect('category_selection')

                except (Element.DoesNotExist, ValueError):
                    messages.error(request, 'Erreur lors de l\'ajout des éléments. Veuillez réessayer.')
            else:
                messages.warning(request, 'Aucun élément sélectionné.')

    return render(request, 'client/item_selection.html', context)

# estimation/views.py
from django.shortcuts import render, get_object_or_404
from django.db.models import Prefetch
from .models import Projet, EstimationElement, DemandeElement, EstimationSummary


def rapport_projet(request, projet_id):
    """Génération du rapport final incluant les demandes personnalisées approuvées et le sablage"""
    projet = get_object_or_404(Projet, id=projet_id)

    # --- Résumé des coûts : on (re)calcule TOUJOURS avant d'afficher ---
    summary, _ = EstimationSummary.objects.get_or_create(projet=projet)  # Correction de la syntaxe

    # Si tu veux permettre d'ajuster la TVA côté URL ?tva=18.0
    tva_param = request.GET.get('tva')
    if tva_param:
        try:
            summary.tva_taux = float(tva_param)
        except ValueError:
            pass  # ignore, on garde le taux existant

    summary.calculer_totaux()

    # --- Données d'affichage ---
    # Éléments standards sélectionnés
    elements_estimation = (
        EstimationElement.objects
        .filter(projet=projet)
        .select_related('element', 'element__categorie', 'element__discipline')
    )

    # Demandes personnalisées approuvées
    demandes_approuvees = (
        DemandeElement.objects
        .filter(
            projet=projet,
            statut='approuve',
            prix_unitaire_admin__isnull=False
        )
        .select_related('categorie', 'discipline')
    )

    # Récupérer les éléments de sablage depuis les sessions utilisateur
    # (Alternative 1: depuis la session Django - pour les sessions en cours)
    elements_sablage_session = []

    # On peut récupérer les sessions de tous les utilisateurs pour ce projet
    # ou implémenter un système de stockage persistant des calculs de sablage

    # Alternative 2: Identifier les éléments de sablage dans EstimationElement
    # (éléments avec element=None et des caractéristiques spécifiques au sablage)
    elements_sablage_estimation = elements_estimation.filter(
        element__isnull=True,  # Pas d'élément standard associé
        # On peut ajouter d'autres critères pour identifier le sablage
        # Par exemple, si on stocke un flag dans les caractéristiques
    )

    # Alternative 3: Utiliser les modèles dédiés (si implémentés)
    sessions_sablage = SessionSablage.objects.filter(projet=projet, valide=True)

    # Grouper par catégorie (éléments standards + demandes approuvées + sablage)
    elements_par_categorie = {}

    # Traiter les éléments standards (non-sablage)
    for elem in elements_estimation:
        if elem.element and elem.element.categorie:
            cat_nom = elem.element.categorie.nom
            if cat_nom not in elements_par_categorie:
                elements_par_categorie[cat_nom] = {
                    'elements': [],
                    'demandes_approuvees': [],
                    'elements_sablage': [],
                    'total': 0,
                    'categorie': elem.element.categorie
                }
            elements_par_categorie[cat_nom]['elements'].append(elem)
            elements_par_categorie[cat_nom]['total'] += elem.cout_total

    # Traiter les éléments de sablage (EstimationElement sans element standard)
    for elem_sablage in elements_sablage_estimation:
        # Déterminer la catégorie (probablement "Main d'œuvre Tuyauterie")
        cat_nom = "Main d'œuvre Tuyauterie"  # ou récupérer depuis un autre champ

        if cat_nom not in elements_par_categorie:
            # Récupérer la catégorie main d'œuvre tuyauterie
            try:
                categorie_mo = Categorie.objects.get(
                    nom__icontains="tuyauterie",
                    type_categorie="main_oeuvre"
                )
            except Categorie.DoesNotExist:
                # Fallback: première catégorie main d'œuvre
                categorie_mo = Categorie.objects.filter(type_categorie="main_oeuvre").first()

            elements_par_categorie[cat_nom] = {
                'elements': [],
                'demandes_approuvees': [],
                'elements_sablage': [],
                'total': 0,
                'categorie': categorie_mo
            }

        elements_par_categorie[cat_nom]['elements_sablage'].append(elem_sablage)
        elements_par_categorie[cat_nom]['total'] += elem_sablage.cout_total

    # Traiter les demandes personnalisées approuvées
    for demande in demandes_approuvees:
        if demande.categorie:
            cat_nom = demande.categorie.nom
            if cat_nom not in elements_par_categorie:
                elements_par_categorie[cat_nom] = {
                    'elements': [],
                    'demandes_approuvees': [],
                    'elements_sablage': [],
                    'total': 0,
                    'categorie': demande.categorie
                }
            elements_par_categorie[cat_nom]['demandes_approuvees'].append(demande)
            elements_par_categorie[cat_nom]['total'] += demande.cout_total

    # Si on utilise les sessions de sablage dédiées (Alternative 3)

    # Décommenter si vous implémentez SessionSablage
    for session in sessions_sablage:
        cat_nom = "Main d'œuvre Tuyauterie"

        if cat_nom not in elements_par_categorie:
            try:
                categorie_mo = Categorie.objects.get(
                    nom__icontains="tuyauterie",
                    type_categorie="main_oeuvre"
                )
            except Categorie.DoesNotExist:
                categorie_mo = Categorie.objects.filter(type_categorie="main_oeuvre").first()

            elements_par_categorie[cat_nom] = {
                'elements': [],
                'demandes_approuvees': [],
                'elements_sablage': [],
                'sessions_sablage': [],
                'total': 0,
                'categorie': categorie_mo
            }

        if 'sessions_sablage' not in elements_par_categorie[cat_nom]:
            elements_par_categorie[cat_nom]['sessions_sablage'] = []

        elements_par_categorie[cat_nom]['sessions_sablage'].append(session)
        elements_par_categorie[cat_nom]['total'] += session.cout_total


    # Récupérer les éléments de sablage depuis les sessions actives
    # (pour afficher les calculs en cours même non validés)
    if hasattr(request, 'session') and 'elements_sablage' in request.session:
        elements_sablage_temp = request.session.get('elements_sablage', [])

        if elements_sablage_temp:
            # Calculer le total temporaire
            PRIX_SABLAGE_M2 = 5000  # Ou récupérer depuis la configuration
            surface_globale_temp = sum(elem['surface_totale'] for elem in elements_sablage_temp)
            prix_total_temp = surface_globale_temp * PRIX_SABLAGE_M2

            cat_nom = "Main d'œuvre Tuyauterie"

            if cat_nom not in elements_par_categorie:
                try:
                    categorie_mo = Categorie.objects.get(
                        nom__icontains="tuyauterie",
                        type_categorie="main_oeuvre"
                    )
                except Categorie.DoesNotExist:
                    categorie_mo = Categorie.objects.filter(type_categorie="main_oeuvre").first()

                elements_par_categorie[cat_nom] = {
                    'elements': [],
                    'demandes_approuvees': [],
                    'elements_sablage': [],
                    'sablage_temporaire': None,
                    'total': 0,
                    'categorie': categorie_mo
                }

            # Ajouter les informations temporaires de sablage
            elements_par_categorie[cat_nom]['sablage_temporaire'] = {
                'elements': elements_sablage_temp,
                'surface_globale': surface_globale_temp,
                'prix_unitaire': PRIX_SABLAGE_M2,
                'prix_total': prix_total_temp,
                'nb_elements': len(elements_sablage_temp)
            }
            # Note: on n'ajoute pas au total car c'est temporaire

    # Nombre total d'éléments (standards + demandes approuvées + sablage)
    total_elements = (
            elements_estimation.count() +
            demandes_approuvees.count()
    )

    # Ajouter les éléments de sablage au compteur s'il y en a
    for data in elements_par_categorie.values():
        if 'elements_sablage' in data:
            total_elements += len(data['elements_sablage'])
        # Si on utilise sessions_sablage:
        if 'sessions_sablage' in data:
            total_elements += len(data['sessions_sablage'])

    # Informations supplémentaires pour le template
    # Calculer les totaux par type pour affichage
    totaux_par_type = {
        'materiel': 0,
        'main_oeuvre': 0,
        'transport': 0,
        'etude': 0,
        'sablage': 0  # Nouveau : total spécifique au sablage
    }

    for cat_nom, data in elements_par_categorie.items():
        if data['categorie']:
            type_cat = data['categorie'].type_categorie
            if type_cat in totaux_par_type:
                totaux_par_type[type_cat] += data['total']

            # Calculer spécifiquement les totaux de sablage
            if 'elements_sablage' in data and data['elements_sablage']:
                totaux_par_type['sablage'] += sum(elem.cout_total for elem in data['elements_sablage'])

            # Si temporaire
            if 'sablage_temporaire' in data and data['sablage_temporaire']:
                totaux_par_type['sablage'] += data['sablage_temporaire']['prix_total']

    context = {
        'projet': projet,
        'elements_par_categorie': elements_par_categorie,
        'summary': summary,
        'total_elements': total_elements,
        'totaux_par_type': totaux_par_type,  # Nouveau pour affichage détaillé
        'has_sablage': any(
            'elements_sablage' in data and data['elements_sablage'] or
            'sablage_temporaire' in data and data['sablage_temporaire']
            for data in elements_par_categorie.values()
        ),  # Flag pour conditions dans le template
    }

    return render(request, 'client/rapport.html', context)
def ajax_update_quantity(request):
    """Mise à jour AJAX des quantités"""
    if request.method == 'POST':
        data = json.loads(request.body)
        element_id = data.get('element_id')
        nouvelle_quantite = data.get('quantite')
        projet_id = request.session.get('projet_id')

        if element_id and nouvelle_quantite and projet_id:
            try:
                estimation_element = EstimationElement.objects.get(
                    element_id=element_id,
                    projet_id=projet_id
                )
                estimation_element.quantite = float(nouvelle_quantite)
                estimation_element.save()

                # Recalculer le résumé
                summary = EstimationSummary.objects.get(projet_id=projet_id)
                summary.calculer_totaux()

                return JsonResponse({
                    'success': True,
                    'nouveau_total': float(estimation_element.cout_total)
                })
            except (EstimationElement.DoesNotExist, ValueError):
                return JsonResponse({'success': False})

    return JsonResponse({'success': False})


# Ajoutez ces vues à votre fichier estimation/views.py

def demandes_personnalisees(request):
    """Affichage des demandes personnalisées du client"""
    projet_id = request.session.get('projet_id')
    if not projet_id:
        return redirect('project_selection')

    projet = get_object_or_404(Projet, id=projet_id)

    # Récupérer toutes les demandes d'éléments personnalisés pour ce projet
    demandes = DemandeElement.objects.filter(
        projet=projet
    ).select_related('categorie', 'discipline').order_by('-date_demande')

    context = {
        'projet': projet,
        'demandes': demandes,
    }

    return render(request, 'client/demandes_personnalisees.html', context)


def supprimer_demande(request, demande_id):
    """Suppression d'une demande d'élément personnalisé"""
    projet_id = request.session.get('projet_id')
    if not projet_id:
        return redirect('project_selection')

    projet = get_object_or_404(Projet, id=projet_id)

    # Récupérer la demande et vérifier qu'elle appartient au bon projet
    demande = get_object_or_404(DemandeElement, id=demande_id, projet=projet)

    # On ne peut supprimer que les demandes en attente
    if demande.statut != 'en_attente':
        messages.error(request, 'Vous ne pouvez supprimer que les demandes en attente.')
        return redirect('demandes_personnalisees')

    # Supprimer la demande
    designation = demande.designation
    demande.delete()

    messages.success(request, f'La demande "{designation}" a été supprimée avec succès.')
    return redirect('demandes_personnalisees')


# estimation/views.py - Fonctions d'export mises à jour avec logo

import os
from django.conf import settings
from reportlab.lib.utils import ImageReader
from reportlab.platypus import Image, Paragraph, Table, TableStyle, Spacer, SimpleDocTemplate
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import datetime


# estimation/views.py - Corrections pour inclure le sablage temporaire dans les exports

def export_pdf_reportlab(request, projet_id):
    """Export PDF avec sablage + colonnes Caractéristiques & Unité adaptées"""
    # --- Imports locaux (auto-contenu) ---
    import os, re, datetime
    from io import BytesIO
    from decimal import Decimal
    from xml.sax.saxutils import escape as xml_escape

    from django.conf import settings
    from django.shortcuts import get_object_or_404
    from django.http import HttpResponse

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import inch
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Table, TableStyle, Image, Spacer
    )

    # --- Données projet ---
    projet = get_object_or_404(Projet, id=projet_id)

    def format_caracteristiques(text, max_length=80):
        if not text or len(text) <= max_length:
            return text or '-'
        truncated = text[:max_length - 3].rsplit(' ', 1)[0] + '...'
        return truncated

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Heading1'],
        fontSize=18, spaceAfter=30, alignment=1, textColor=colors.HexColor('#667eea')
    )
    heading_style = ParagraphStyle(
        'CustomHeading', parent=styles['Heading2'],
        fontSize=14, spaceAfter=12, textColor=colors.HexColor('#333333')
    )
    carac_para_style = ParagraphStyle(
        'CaracCell', parent=styles['Normal'],
        fontSize=9, leading=12, spaceBefore=0, spaceAfter=0
    )
    unit_para_style = ParagraphStyle(  # <<< NOUVEAU : pour la colonne Unité
        'UnitCell', parent=styles['Normal'],
        fontSize=9, leading=11, alignment=1  # 1 = center
    )

    def caracs_paragraph(text: str) -> Paragraph:
        if not text:
            return Paragraph('-', carac_para_style)
        parts = re.split(r'\s*-\s*|;\s*', text.strip())
        pretty = '<br/>'.join(f'• {xml_escape(p)}' for p in parts if p)
        return Paragraph(pretty, carac_para_style)

    def unit_paragraph(text: str) -> Paragraph:
        return Paragraph(xml_escape(text or '-'), unit_para_style)

    # --- Récupération données ---
    elements_selections = EstimationElement.objects.filter(projet=projet).select_related(
        'element', 'element__categorie', 'element__discipline'
    )
    demandes_approuvees = DemandeElement.objects.filter(projet=projet, statut='approuve')

    elements_sablage_temp = request.session.get('elements_sablage', [])
    PRIX_SABLAGE_M2 = 5000

    elements_par_categorie = {}

    for selection in elements_selections:
        if selection.element:
            cat_nom = selection.element.categorie.nom
            if cat_nom not in elements_par_categorie:
                elements_par_categorie[cat_nom] = {
                    'categorie': selection.element.categorie,
                    'elements': [], 'demandes_approuvees': [],
                    'elements_sablage': [], 'total': 0
                }
            elements_par_categorie[cat_nom]['elements'].append(selection)
            elements_par_categorie[cat_nom]['total'] += selection.cout_total
        else:
            cat_nom = "Main d'œuvre Tuyauterie"
            if cat_nom not in elements_par_categorie:
                try:
                    categorie_mo = Categorie.objects.filter(type_categorie="main_oeuvre").first()
                except Categorie.DoesNotExist:
                    categorie_mo = type('TempCategorie', (), {
                        'nom': "Main d'œuvre Tuyauterie", 'type_categorie': 'main_oeuvre'
                    })()
                elements_par_categorie[cat_nom] = {
                    'categorie': categorie_mo,
                    'elements': [], 'demandes_approuvees': [],
                    'elements_sablage': [], 'total': 0
                }
            elements_par_categorie[cat_nom]['elements_sablage'].append(selection)
            elements_par_categorie[cat_nom]['total'] += selection.cout_total

    if elements_sablage_temp:
        cat_nom = "Main d'œuvre Tuyauterie"
        if cat_nom not in elements_par_categorie:
            try:
                categorie_mo = Categorie.objects.filter(type_categorie="main_oeuvre").first()
            except Categorie.DoesNotExist:
                categorie_mo = type('TempCategorie', (), {
                    'nom': "Main d'œuvre Tuyauterie", 'type_categorie': 'main_oeuvre'
                })()
            elements_par_categorie[cat_nom] = {
                'categorie': categorie_mo,
                'elements': [], 'demandes_approuvees': [],
                'elements_sablage': [], 'elements_sablage_temp': [], 'total': 0
            }

        surface_globale_temp = sum(elem['surface_totale'] for elem in elements_sablage_temp)
        prix_total_temp = Decimal(str(surface_globale_temp * PRIX_SABLAGE_M2))
        elements_par_categorie[cat_nom]['elements_sablage_temp'] = {
            'elements': elements_sablage_temp,
            'surface_globale': surface_globale_temp,
            'prix_unitaire': PRIX_SABLAGE_M2,
            'prix_total': prix_total_temp,
            'nb_elements': len(elements_sablage_temp)
        }
        elements_par_categorie[cat_nom]['total'] += prix_total_temp

    for demande in demandes_approuvees:
        cat_nom = demande.categorie.nom
        if cat_nom not in elements_par_categorie:
            elements_par_categorie[cat_nom] = {
                'categorie': demande.categorie,
                'elements': [], 'demandes_approuvees': [],
                'elements_sablage': [], 'total': 0
            }
        elements_par_categorie[cat_nom]['demandes_approuvees'].append(demande)
        elements_par_categorie[cat_nom]['total'] += demande.cout_total

    try:
        summary = EstimationSummary.objects.get(projet=projet)
        summary.calculer_totaux()
        if elements_sablage_temp:
            surface_globale_temp = sum(elem['surface_totale'] for elem in elements_sablage_temp)
            prix_sablage_temp = Decimal(str(surface_globale_temp * PRIX_SABLAGE_M2))
            summary.cout_total_main_oeuvre += prix_sablage_temp
            summary.cout_total_ht += prix_sablage_temp
            summary.tva_montant = summary.cout_total_ht * (summary.tva_taux / Decimal('100'))
            summary.cout_total_ttc = summary.cout_total_ht + summary.tva_montant
    except EstimationSummary.DoesNotExist:
        summary = EstimationSummary.objects.create(projet=projet)
        summary.calculer_totaux()

    # --- PDF ---
    response = HttpResponse(content_type='application/pdf')
    nom_fichier_securise = "".join(c for c in projet.nom if c.isalnum() or c in (' ', '-', '_')).rstrip()
    response['Content-Disposition'] = f'attachment; filename="rapport_{nom_fichier_securise}_{datetime.date.today()}.pdf"'

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    story = []

    logo_path = os.path.join(settings.STATIC_ROOT or settings.BASE_DIR, 'static', 'img', 'logo.jpg')
    if os.path.exists(logo_path):
        try:
            logo_img = Image(logo_path, width=2 * inch, height=1 * inch)
            logo_img.hAlign = 'CENTER'
            story.append(logo_img)
            story.append(Spacer(1, 20))
        except Exception:
            pass

    story.append(Paragraph("<b>VOTRE ENTREPRISE</b><br/>Adresse de l'entreprise<br/>Téléphone - Email", styles['Normal']))
    story.append(Spacer(1, 20))
    story.append(Paragraph(f"RAPPORT D'ESTIMATION<br/>Projet: {projet.nom}", title_style))
    if projet.client:
        story.append(Paragraph(f"Client: <b>{projet.client}</b>", styles['Normal']))
    story.append(Paragraph(f"Date: {datetime.date.today().strftime('%d/%m/%Y')}", styles['Normal']))
    story.append(Spacer(1, 20))

    # --- Tableaux par catégorie ---
    for categorie_nom, data in elements_par_categorie.items():
        story.append(Paragraph(f"{categorie_nom} - Total: {data['total']:,.2f} CFA", heading_style))

        table_data = [['Désignation', 'Caractéristiques', 'Prix Unit.', 'Qté', 'Unité', 'Total']]

        # Éléments standards
        for element in data['elements']:
            table_data.append([
                element.designation,
                caracs_paragraph(element.caracteristiques),
                f"{element.prix_unitaire_utilise:,.2f}",
                f"{element.quantite:,.2f}",
                unit_paragraph(element.unite_display),    # <<< Paragraph + wrap
                f"{element.cout_total:,.2f}"
            ])

        # Sablage validé
        for element_sablage in data['elements_sablage']:
            sablage_para = Paragraph(xml_escape(f"Surface: {element_sablage.quantite:.3f} m²"), carac_para_style)
            table_data.append([
                "Sablage Tuyauterie",
                sablage_para,
                f"{element_sablage.prix_unitaire_utilise:,.2f}",
                f"{element_sablage.quantite:,.3f}",
                unit_paragraph("m²"),
                f"{element_sablage.cout_total:,.2f}"
            ])

        # Sablage temporaire
        if data.get('elements_sablage_temp'):
            sablage_temp = data['elements_sablage_temp']
            elements_resume = [f"{e['nom_type_piece']} {e['nom_dn']} ({e['quantite']:g})" for e in sablage_temp['elements']]
            resume_text = ", ".join(elements_resume)
            if len(resume_text) > 80:
                words = resume_text.split(', ')
                lines, cur = [], ""
                for w in words:
                    if len(cur + w) <= 80: cur += ("" if not cur else ", ") + w
                    else: lines.append(cur); cur = w
                if cur: lines.append(cur)
                resume_final = "<br/>".join(xml_escape(l) for l in lines)
            else:
                resume_final = xml_escape(resume_text)

            sablage_temp_para = Paragraph(
                f"Surface: {sablage_temp['surface_globale']:.3f} m²<br/>Détail: {resume_final}",
                carac_para_style
            )
            table_data.append([
                "Sablage Tuyauterie",
                sablage_temp_para,
                f"{sablage_temp['prix_unitaire']:,.2f}",
                f"{sablage_temp['surface_globale']:,.3f}",
                unit_paragraph("m²"),
                f"{sablage_temp['prix_total']:,.2f}"
            ])

        # Demandes perso
        for demande in data['demandes_approuvees']:
            table_data.append([
                f"{demande.designation} (Personnalisé)",
                caracs_paragraph(demande.caracteristiques),
                f"{demande.prix_unitaire_admin:,.2f}",
                f"{demande.quantite:,.2f}",
                unit_paragraph(demande.get_unite_display()),
                f"{demande.cout_total:,.2f}"
            ])

        # Largeurs : Unité plus large (évite le débordement)
        is_materiel = (
            getattr(data['categorie'], 'type_categorie', '') == 'materiel' or
            'Matériel' in categorie_nom or 'Materiel' in categorie_nom
        )
        has_sablage = any('Sablage' in str(r[0]) for r in table_data[1:])

        if is_materiel:
            colWidths = [1.6 * inch, 3.3 * inch, 0.85 * inch, 0.60 * inch, 1.05 * inch, 0.85 * inch]
        elif has_sablage:
            colWidths = [1.6 * inch, 3.0 * inch, 0.80 * inch, 0.55 * inch, 0.90 * inch, 0.85 * inch]
        else:
            colWidths = [1.8 * inch, 2.7 * inch, 0.80 * inch, 0.55 * inch, 0.90 * inch, 0.75 * inch]

        table = Table(table_data, colWidths=colWidths)

        # Styles
        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),

            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),

            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),

            # Alignements COLONNE par COLONNE (plus de "tout à droite")
            ('ALIGN', (0, 0), (1, -1), 'LEFT'),   # Désignation, Caractéristiques
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),  # Prix Unit.
            ('ALIGN', (3, 0), (3, -1), 'RIGHT'),  # Qté
            ('ALIGN', (4, 0), (4, -1), 'CENTER'), # Unité
            ('ALIGN', (5, 0), (5, -1), 'RIGHT'),  # Total

            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('WORDWRAP', (0, 0), (-1, -1), 'LTR'),
        ]

        # Épaisseur supplémentaire pour Caractéristiques des Matériels
        if is_materiel:
            table_style.extend([
                ('TOPPADDING',    (1, 1), (1, -1), 8),
                ('BOTTOMPADDING', (1, 1), (1, -1), 8),
                ('LEFTPADDING',   (1, 1), (1, -1), 6),
                ('RIGHTPADDING',  (1, 1), (1, -1), 6),
                ('FONTSIZE',      (1, 1), (1, -1), 9),
            ])

        # Style Sablage
        for i, row in enumerate(table_data[1:], 1):
            if 'Sablage' in str(row[0]):
                table_style.extend([
                    ('FONTSIZE', (0, i), (-1, i), 7),
                    ('TOPPADDING', (0, i), (-1, i), 8),
                    ('BOTTOMPADDING', (0, i), (-1, i), 8),
                    ('BACKGROUND', (0, i), (-1, i), colors.HexColor('#FFF3CD')),
                ])

        table.setStyle(TableStyle(table_style))
        story.append(table)
        story.append(Spacer(1, 20))

    # --- Résumé financier ---
    story.append(Paragraph("RÉSUMÉ FINANCIER", heading_style))
    financial_data = [
        ['Sous-total HT', f"{summary.cout_total_ht:,.2f} CFA"],
        [f"TVA ({summary.tva_taux}%)", f"{summary.tva_montant:,.2f} CFA"],
        ['TOTAL TTC', f"{summary.cout_total_ttc:,.2f} CFA"],
    ]
    financial_table = Table(financial_data, colWidths=[3 * inch, 2 * inch])
    financial_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    story.append(financial_table)

    if elements_sablage_temp:
        story.append(Spacer(1, 15))
        note_sablage = Paragraph(
            f"<i>Note : Ce rapport inclut {len(elements_sablage_temp)} élément(s) de sablage temporaire pour un total de "
            f"{sum(elem['surface_totale'] for elem in elements_sablage_temp):.3f} m²</i>",
            styles['Normal']
        )
        story.append(note_sablage)

    story.append(Spacer(1, 30))
    story.append(Paragraph(
        f"<i>Rapport d'estimation généré pour le projet \"{projet.nom}\" le "
        f"{datetime.date.today().strftime('%d/%m/%Y')}</i>", styles['Normal']
    ))

    # Build & réponse
    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response


def export_excel_advanced(request, projet_id):
    """Export Excel avec sablage temporaire inclus"""
    projet = get_object_or_404(Projet, id=projet_id)

    # Récupérer les données
    elements_selections = EstimationElement.objects.filter(projet=projet).select_related(
        'element', 'element__categorie', 'element__discipline'
    )
    demandes_approuvees = DemandeElement.objects.filter(projet=projet, statut='approuve')

    # NOUVEAU : Récupérer le sablage temporaire
    elements_sablage_temp = request.session.get('elements_sablage', [])
    PRIX_SABLAGE_M2 = 5000

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rapport d'Estimation"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    title_font = Font(bold=True, size=16, color="667EEA")
    category_font = Font(bold=True, size=14, color="333333")
    company_font = Font(bold=True, size=12, color="333333")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center')
    right_alignment = Alignment(horizontal='right', vertical='center')

    row = 1

    # Logo
    logo_path = os.path.join(settings.STATIC_ROOT or settings.BASE_DIR, 'static', 'img', 'logo.jpg')
    if os.path.exists(logo_path):
        try:
            with PILImage.open(logo_path) as pil_img:
                if pil_img.height > 150:
                    ratio = 150 / pil_img.height
                    new_width = int(pil_img.width * ratio)
                    pil_img = pil_img.resize((new_width, 150), PILImage.Resampling.LANCZOS)

                img_buffer = BytesIO()
                pil_img.save(img_buffer, format='PNG')
                img_buffer.seek(0)

                excel_img = ExcelImage(img_buffer)
                excel_img.anchor = f'A{row}'
                ws.add_image(excel_img)

                for i in range(row, row + 8):
                    ws.row_dimensions[i].height = 20
                row += 8
        except Exception as e:
            print(f"Erreur lors du chargement du logo dans Excel: {e}")

    # En-tête
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "VOTRE ENTREPRISE"
    ws[f'A{row}'].font = company_font
    ws[f'A{row}'].alignment = center_alignment
    row += 1

    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "Adresse de l'entreprise - Téléphone - Email"
    ws[f'A{row}'].font = Font(size=10)
    ws[f'A{row}'].alignment = center_alignment
    row += 2

    # Titre
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = f"RAPPORT D'ESTIMATION - Projet: {projet.nom}"
    ws[f'A{row}'].font = title_font
    ws[f'A{row}'].alignment = center_alignment
    row += 2

    # Infos projet
    if projet.client:
        ws[f'A{row}'] = f"Client: {projet.client}"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

    ws[f'A{row}'] = f"Date: {datetime.date.today().strftime('%d/%m/%Y')}"
    ws[f'A{row}'].font = Font(bold=True)
    row += 2

    # Organiser par catégorie (même logique que PDF)
    elements_par_categorie = {}

    # Traiter éléments standards et sablage validé
    for selection in elements_selections:
        if selection.element:
            cat_nom = selection.element.categorie.nom
            if cat_nom not in elements_par_categorie:
                elements_par_categorie[cat_nom] = {
                    'elements': [], 'demandes_approuvees': [], 'elements_sablage': [], 'total': 0
                }
            elements_par_categorie[cat_nom]['elements'].append(selection)
            elements_par_categorie[cat_nom]['total'] += selection.cout_total
        else:
            cat_nom = "Main d'œuvre Tuyauterie"
            if cat_nom not in elements_par_categorie:
                elements_par_categorie[cat_nom] = {
                    'elements': [], 'demandes_approuvees': [], 'elements_sablage': [], 'total': 0
                }
            elements_par_categorie[cat_nom]['elements_sablage'].append(selection)
            elements_par_categorie[cat_nom]['total'] += selection.cout_total

    # NOUVEAU : Ajouter sablage temporaire
    if elements_sablage_temp:
        from decimal import Decimal
        cat_nom = "Main d'œuvre Tuyauterie"
        if cat_nom not in elements_par_categorie:
            elements_par_categorie[cat_nom] = {
                'elements': [], 'demandes_approuvees': [], 'elements_sablage': [],
                'elements_sablage_temp': [], 'total': Decimal('0')
            }

        surface_globale_temp = sum(elem['surface_totale'] for elem in elements_sablage_temp)
        prix_total_temp = Decimal(str(surface_globale_temp * PRIX_SABLAGE_M2))

        elements_par_categorie[cat_nom]['elements_sablage_temp'] = {
            'elements': elements_sablage_temp,
            'surface_globale': surface_globale_temp,
            'prix_unitaire': PRIX_SABLAGE_M2,
            'prix_total': prix_total_temp,
            'nb_elements': len(elements_sablage_temp)
        }
        elements_par_categorie[cat_nom]['total'] += prix_total_temp

    # Ajouter demandes personnalisées
    for demande in demandes_approuvees:
        cat_nom = demande.categorie.nom
        if cat_nom not in elements_par_categorie:
            elements_par_categorie[cat_nom] = {
                'elements': [], 'demandes_approuvees': [], 'elements_sablage': [], 'total': 0
            }
        elements_par_categorie[cat_nom]['demandes_approuvees'].append(demande)
        elements_par_categorie[cat_nom]['total'] += demande.cout_total

    # Données par catégorie
    for categorie_nom, data in elements_par_categorie.items():
        # Titre catégorie
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = f"{categorie_nom} - Total: {data['total']:,.2f} CFA"
        ws[f'A{row}'].font = category_font
        ws[f'A{row}'].alignment = center_alignment
        row += 1

        # En-têtes
        headers = ['Désignation', 'Caractéristiques', 'Prix Unitaire', 'Quantité', 'Unité', 'Total']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        row += 1

        # Éléments standards
        for element in data['elements']:
            ws.cell(row=row, column=1, value=element.designation).border = border
            ws.cell(row=row, column=2, value=element.caracteristiques or '-').border = border

            prix_cell = ws.cell(row=row, column=3, value=float(element.prix_unitaire_utilise))
            prix_cell.number_format = '#,##0.00'
            prix_cell.alignment = right_alignment
            prix_cell.border = border

            qte_cell = ws.cell(row=row, column=4, value=float(element.quantite))
            qte_cell.number_format = '#,##0.00'
            qte_cell.alignment = right_alignment
            qte_cell.border = border

            ws.cell(row=row, column=5, value=element.unite_display).border = border

            total_cell = ws.cell(row=row, column=6, value=float(element.cout_total))
            total_cell.number_format = '#,##0.00'
            total_cell.alignment = right_alignment
            total_cell.border = border
            total_cell.font = Font(bold=True)

            row += 1

        # Éléments sablage validés
        for element_sablage in data['elements_sablage']:
            sablage_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

            designation_cell = ws.cell(row=row, column=1, value="Sablage Tuyauterie")
            designation_cell.border = border
            designation_cell.fill = sablage_fill

            carac_cell = ws.cell(row=row, column=2, value=f"Surface: {element_sablage.quantite:.3f} m²")
            carac_cell.border = border
            carac_cell.fill = sablage_fill

            prix_cell = ws.cell(row=row, column=3, value=float(element_sablage.prix_unitaire_utilise))
            prix_cell.number_format = '#,##0.00'
            prix_cell.alignment = right_alignment
            prix_cell.border = border
            prix_cell.fill = sablage_fill

            qte_cell = ws.cell(row=row, column=4, value=float(element_sablage.quantite))
            qte_cell.number_format = '#,##0.000'
            qte_cell.alignment = right_alignment
            qte_cell.border = border
            qte_cell.fill = sablage_fill

            unite_cell = ws.cell(row=row, column=5, value="m²")
            unite_cell.border = border
            unite_cell.fill = sablage_fill

            total_cell = ws.cell(row=row, column=6, value=float(element_sablage.cout_total))
            total_cell.number_format = '#,##0.00'
            total_cell.alignment = right_alignment
            total_cell.border = border
            total_cell.font = Font(bold=True)
            total_cell.fill = sablage_fill

            row += 1

        # NOUVEAU : Sablage temporaire
        if 'elements_sablage_temp' in data and data['elements_sablage_temp']:
            sablage_temp = data['elements_sablage_temp']
            sablage_temp_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")

            # Créer résumé
            elements_resume = []
            for elem in sablage_temp['elements']:
                elements_resume.append(f"{elem['nom_type_piece']} {elem['nom_dn']} ({elem['quantite']:g})")
            resume_text = ", ".join(elements_resume)

            designation_cell = ws.cell(row=row, column=1, value="Sablage Tuyauterie (Temporaire)")
            designation_cell.border = border
            designation_cell.fill = sablage_temp_fill

            carac_cell = ws.cell(row=row, column=2,
                                 value=f"Surface: {sablage_temp['surface_globale']:.3f} m² - {resume_text}")
            carac_cell.border = border
            carac_cell.fill = sablage_temp_fill

            prix_cell = ws.cell(row=row, column=3, value=float(sablage_temp['prix_unitaire']))
            prix_cell.number_format = '#,##0.00'
            prix_cell.alignment = right_alignment
            prix_cell.border = border
            prix_cell.fill = sablage_temp_fill

            qte_cell = ws.cell(row=row, column=4, value=float(sablage_temp['surface_globale']))
            qte_cell.number_format = '#,##0.000'
            qte_cell.alignment = right_alignment
            qte_cell.border = border
            qte_cell.fill = sablage_temp_fill

            unite_cell = ws.cell(row=row, column=5, value="m²")
            unite_cell.border = border
            unite_cell.fill = sablage_temp_fill

            total_cell = ws.cell(row=row, column=6, value=float(sablage_temp['prix_total']))
            total_cell.number_format = '#,##0.00'
            total_cell.alignment = right_alignment
            total_cell.border = border
            total_cell.font = Font(bold=True)
            total_cell.fill = sablage_temp_fill

            row += 1

        # Éléments personnalisés
        for demande in data['demandes_approuvees']:
            perso_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")

            designation_cell = ws.cell(row=row, column=1, value=f"{demande.designation} (Personnalisé)")
            designation_cell.border = border
            designation_cell.fill = perso_fill

            carac_cell = ws.cell(row=row, column=2, value=demande.caracteristiques or '-')
            carac_cell.border = border
            carac_cell.fill = perso_fill

            prix_cell = ws.cell(row=row, column=3, value=float(demande.prix_unitaire_admin))
            prix_cell.number_format = '#,##0.00'
            prix_cell.alignment = right_alignment
            prix_cell.border = border
            prix_cell.fill = perso_fill

            qte_cell = ws.cell(row=row, column=4, value=float(demande.quantite))
            qte_cell.number_format = '#,##0.00'
            qte_cell.alignment = right_alignment
            qte_cell.border = border
            qte_cell.fill = perso_fill

            unite_cell = ws.cell(row=row, column=5, value=demande.get_unite_display())
            unite_cell.border = border
            unite_cell.fill = perso_fill

            total_cell = ws.cell(row=row, column=6, value=float(demande.cout_total))
            total_cell.number_format = '#,##0.00'
            total_cell.alignment = right_alignment
            total_cell.border = border
            total_cell.font = Font(bold=True)
            total_cell.fill = perso_fill

            row += 1

        row += 1

    # Calcul du résumé incluant sablage temporaire
    try:
        summary = EstimationSummary.objects.get(projet=projet)
        summary.calculer_totaux()

        if elements_sablage_temp:
            from decimal import Decimal
            surface_globale_temp = sum(elem['surface_totale'] for elem in elements_sablage_temp)
            prix_sablage_temp = Decimal(str(surface_globale_temp * PRIX_SABLAGE_M2))
            summary.cout_total_main_oeuvre += prix_sablage_temp
            summary.cout_total_ht += prix_sablage_temp
            summary.tva_montant = summary.cout_total_ht * (summary.tva_taux / Decimal('100'))
            summary.cout_total_ttc = summary.cout_total_ht + summary.tva_montant

    except EstimationSummary.DoesNotExist:
        summary = EstimationSummary.objects.create(projet=projet)
        summary.calculer_totaux()

    # Résumé financier
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "RÉSUMÉ FINANCIER"
    ws[f'A{row}'].font = category_font
    ws[f'A{row}'].alignment = center_alignment
    row += 1

    # Sous-total HT
    ws.cell(row=row, column=4, value="Sous-total HT:").font = Font(bold=True)
    ws.cell(row=row, column=4).alignment = right_alignment
    total_ht_cell = ws.cell(row=row, column=6, value=float(summary.cout_total_ht))
    total_ht_cell.number_format = '#,##0.00 "CFA"'
    total_ht_cell.alignment = right_alignment
    total_ht_cell.font = Font(bold=True)
    row += 1

    # TVA
    ws.cell(row=row, column=4, value=f"TVA ({summary.tva_taux}%):").font = Font(bold=True)
    ws.cell(row=row, column=4).alignment = right_alignment
    tva_cell = ws.cell(row=row, column=6, value=float(summary.tva_montant))
    tva_cell.number_format = '#,##0.00 "CFA"'
    tva_cell.alignment = right_alignment
    tva_cell.font = Font(bold=True)
    row += 1

    # Total TTC
    ws.cell(row=row, column=4, value="TOTAL TTC:").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row, column=4).fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    ws.cell(row=row, column=4).alignment = right_alignment

    total_ttc_cell = ws.cell(row=row, column=6, value=float(summary.cout_total_ttc))
    total_ttc_cell.number_format = '#,##0.00 "CFA"'
    total_ttc_cell.alignment = right_alignment
    total_ttc_cell.font = Font(bold=True, color="FFFFFF")
    total_ttc_cell.fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")

    # Largeurs colonnes
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 15

    # Note sur sablage temporaire
    if elements_sablage_temp:
        row += 3
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = f"Note: Ce rapport inclut {len(elements_sablage_temp)} élément(s) de sablage temporaire"
        ws[f'A{row}'].font = Font(italic=True, size=10)
        ws[f'A{row}'].alignment = center_alignment

    # Pied de page
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws[
        f'A{row}'] = f"Rapport d'estimation généré pour le projet \"{projet.nom}\" le {datetime.date.today().strftime('%d/%m/%Y')}"
    ws[f'A{row}'].font = Font(italic=True, size=10)
    ws[f'A{row}'].alignment = center_alignment

    # Réponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    nom_fichier_securise = "".join(c for c in projet.nom if c.isalnum() or c in (' ', '-', '_')).rstrip()
    response[
        'Content-Disposition'] = f'attachment; filename="rapport_{nom_fichier_securise}_{datetime.date.today().strftime("%Y%m%d")}.xlsx"'

    wb.save(response)
    return response

##################
# estimation/views.py - Ajouter cette nouvelle vue

def sablage_tuyauterie(request, categorie_id):
    """Gestion du sablage tuyauterie avec calcul de surface"""
    projet_id = request.session.get('projet_id')
    if not projet_id:
        return redirect('project_selection')

    projet = get_object_or_404(Projet, id=projet_id)
    categorie = get_object_or_404(Categorie, id=categorie_id)

    # Données des surfaces unitaires par DN et type de pièce (selon votre tableau Excel)
    SURFACES_SABLAGE = {
        15: {  # DN 15 (1/2")
            'tube': 0.067, 'coude_90': 0.004, 'coude_45': 0.002,
            'coude_90_r5d': 0.007, 'coude_secteur': 0, 'te': 0,
            'bride': 0, 'reduction': 0, 'cap': 0
        },
        20: {  # DN 20 (3/4")
            'tube': 0.084, 'coude_90': 0.005, 'coude_45': 0.003,
            'coude_90_r5d': 0.013, 'coude_secteur': 0, 'te': 0,
            'bride': 0, 'reduction': 0, 'cap': 0
        },
        25: {  # DN 25 (1")
            'tube': 0.105, 'coude_90': 0.006, 'coude_45': 0.003,
            'coude_90_r5d': 0.021, 'coude_secteur': 0, 'te': 0,
            'bride': 0, 'reduction': 0, 'cap': 0
        },
        32: {  # DN 32 (1" 1/4)
            'tube': 0.133, 'coude_90': 0.010, 'coude_45': 0.005,
            'coude_90_r5d': 0.033, 'coude_secteur': 0, 'te': 0,
            'bride': 0, 'reduction': 0, 'cap': 0
        },
        40: {  # DN 40 (1" 1/2)
            'tube': 0.152, 'coude_90': 0.014, 'coude_45': 0.007,
            'coude_90_r5d': 0.045, 'coude_secteur': 0, 'te': 0,
            'bride': 0, 'reduction': 0, 'cap': 0
        },
        50: {  # DN 50 (2")
            'tube': 0.189, 'coude_90': 0.023, 'coude_45': 0.011,
            'coude_90_r5d': 0.076, 'coude_secteur': 0, 'te': 0,
            'bride': 0, 'reduction': 0, 'cap': 0
        },
        65: {  # DN 65 (2" 1/2)
            'tube': 0.229, 'coude_90': 0.034, 'coude_45': 0.017,
            'coude_90_r5d': 0.114, 'coude_secteur': 0, 'te': 0,
            'bride': 0, 'reduction': 0, 'cap': 0
        },
        80: {  # DN 80 (3")
            'tube': 0.279, 'coude_90': 0.050, 'coude_45': 0.025,
            'coude_90_r5d': 0.167, 'coude_secteur': 0, 'te': 0,
            'bride': 0, 'reduction': 0, 'cap': 0
        },
        100: {  # DN 100 (4")
            'tube': 0.359, 'coude_90': 0.086, 'coude_45': 0.043,
            'coude_90_r5d': 0.287, 'coude_secteur': 0, 'te': 0,
            'bride': 0, 'reduction': 0, 'cap': 0
        },
        125: {  # DN 125 (5")
            'tube': 0.439, 'coude_90': 0.131, 'coude_45': 0.066,
            'coude_90_r5d': 0.438, 'coude_secteur': 0, 'te': 0,
            'bride': 0, 'reduction': 0, 'cap': 0
        },
        150: {  # DN 150 (6")
            'tube': 0.529, 'coude_90': 0.190, 'coude_45': 0.095,
            'coude_90_r5d': 0.633, 'coude_secteur': 0.194, 'te': 0,
            'bride': 0, 'reduction': 0, 'cap': 0
        },
        200: {  # DN 200 (8")
            'tube': 0.688, 'coude_90': 0.330, 'coude_45': 0.165,
            'coude_90_r5d': 1.099, 'coude_secteur': 0.338, 'te': 0,
            'bride': 0, 'reduction': 0, 'cap': 0
        },
        250: {  # DN 250 (10")
            'tube': 0.858, 'coude_90': 0.513, 'coude_45': 0.257,
            'coude_90_r5d': 1.711, 'coude_secteur': 0.525, 'te': 0,
            'bride': 0, 'reduction': 0, 'cap': 0
        },
        300: {  # DN 300 (12")
            'tube': 1.018, 'coude_90': 0.731, 'coude_45': 0.365,
            'coude_90_r5d': 2.436, 'coude_secteur': 0.748, 'te': 0,
            'bride': 0, 'reduction': 0, 'cap': 0
        },
        350: {  # DN 350 (14")
            'tube': 1.117, 'coude_90': 0.936, 'coude_45': 0.468,
            'coude_90_r5d': 3.120, 'coude_secteur': 0.958, 'te': 0,
            'bride': 0, 'reduction': 0, 'cap': 0
        },
        400: {  # DN 400 (16")
            'tube': 1.277, 'coude_90': 1.222, 'coude_45': 0.611,
            'coude_90_r5d': 4.076, 'coude_secteur': 1.252, 'te': 0,
            'bride': 0, 'reduction': 0, 'cap': 0
        },
        450: {  # DN 450 (18")
            'tube': 1.436, 'coude_90': 1.547, 'coude_45': 0.774,
            'coude_90_r5d': 5.158, 'coude_secteur': 1.585, 'te': 0,
            'bride': 0, 'reduction': 0, 'cap': 0
        },
        500: {  # DN 500 (20")
            'tube': 1.596, 'coude_90': 1.910, 'coude_45': 0.955,
            'coude_90_r5d': 6.368, 'coude_secteur': 1.953, 'te': 0,
            'bride': 0, 'reduction': 0, 'cap': 0
        },
        600: {  # DN 600 (24")
            'tube': 1.915, 'coude_90': 2.751, 'coude_45': 1.376,
            'coude_90_r5d': 9.169, 'coude_secteur': 2.815, 'te': 0,
            'bride': 0, 'reduction': 0, 'cap': 0
        },
        700: {  # DN 700 (28")
            'tube': 2.234, 'coude_90': 3.744, 'coude_45': 1.872,
            'coude_90_r5d': 12.48, 'coude_secteur': 3.283, 'te': 0,
            'bride': 0, 'reduction': 0, 'cap': 0
        },
        750: {  # DN 750 (30")
            'tube': 0, 'coude_90': 4.298, 'coude_45': 2.149,
            'coude_90_r5d': 14.33, 'coude_secteur': 0, 'te': 0,
            'bride': 0, 'reduction': 0, 'cap': 0
        },
        800: {  # DN 800 (32")
            'tube': 2.554, 'coude_90': 4.892, 'coude_45': 2.446,
            'coude_90_r5d': 16.30, 'coude_secteur': 5.003, 'te': 0,
            'bride': 0, 'reduction': 0, 'cap': 0
        },
        900: {  # DN 900 (36")
            'tube': 2.871, 'coude_90': 6.185, 'coude_45': 3.093,
            'coude_90_r5d': 20.62, 'coude_secteur': 6.330, 'te': 0,
            'bride': 0, 'reduction': 0, 'cap': 0
        },
        1000: {  # DN 1000 (40")
            'tube': 3.192, 'coude_90': 7.641, 'coude_45': 3.821,
            'coude_90_r5d': 25.47, 'coude_secteur': 7.823, 'te': 0,
            'bride': 0, 'reduction': 0, 'cap': 0
        }
    }

    # Types de pièces disponibles
    TYPES_PIECES = [
        ('tube', 'Tube'),
        ('coude_90', 'Coude 90°'),
        ('coude_45', 'Coude 45°'),
        ('te', 'Té'),
        ('bride', 'Bride'),
        ('reduction', 'Réduction'),
        ('cap', 'Cap'),
        ('coude_90_r5d', 'Coude 90° (r=5D)'),
        ('coude_secteur', 'Coude Secteur'),
    ]

    # DN disponibles avec correspondance pouces
    DN_CHOICES = [
        (15, 'DN 15 (1/2")'),
        (20, 'DN 20 (3/4")'),
        (25, 'DN 25 (1")'),
        (32, 'DN 32 (1" 1/4)'),
        (40, 'DN 40 (1" 1/2)'),
        (50, 'DN 50 (2")'),
        (65, 'DN 65 (2" 1/2)'),
        (80, 'DN 80 (3")'),
        (100, 'DN 100 (4")'),
        (125, 'DN 125 (5")'),
        (150, 'DN 150 (6")'),
        (200, 'DN 200 (8")'),
        (250, 'DN 250 (10")'),
        (300, 'DN 300 (12")'),
        (350, 'DN 350 (14")'),
        (400, 'DN 400 (16")'),
        (450, 'DN 450 (18")'),
        (500, 'DN 500 (20")'),
        (600, 'DN 600 (24")'),
        (700, 'DN 700 (28")'),
        (750, 'DN 750 (30")'),
        (800, 'DN 800 (32")'),
        (900, 'DN 900 (36")'),
        (1000, 'DN 1000 (40")'),
    ]

    # Prix au m² pour le sablage (à ajuster selon vos tarifs)
    PRIX_SABLAGE_M2 = 5000  # CFA par m²

    # Récupérer les éléments de sablage déjà en session
    elements_sablage = request.session.get('elements_sablage', [])

    context = {
        'projet': projet,
        'categorie': categorie,
        'types_pieces': TYPES_PIECES,
        'dn_choices': DN_CHOICES,
        'elements_sablage': elements_sablage,
        'prix_m2': PRIX_SABLAGE_M2,
    }

    if request.method == 'POST':
        if 'ajouter_element' in request.POST:
            # Ajouter un nouvel élément de sablage
            type_piece = request.POST.get('type_piece')
            dn = request.POST.get('dn')
            quantite = request.POST.get('quantite')

            if type_piece and dn and quantite:
                try:
                    dn_int = int(dn)
                    quantite_float = float(quantite)

                    # Récupérer la surface unitaire
                    if dn_int in SURFACES_SABLAGE and type_piece in SURFACES_SABLAGE[dn_int]:
                        surface_unitaire = SURFACES_SABLAGE[dn_int][type_piece]

                        if surface_unitaire > 0:  # Vérifier que la surface existe pour ce type/DN
                            surface_totale = quantite_float * surface_unitaire

                            # Obtenir le nom lisible du type de pièce
                            nom_type_piece = dict(TYPES_PIECES).get(type_piece, type_piece)
                            nom_dn = dict(DN_CHOICES).get(dn_int, f'DN {dn_int}')

                            nouvel_element = {
                                'type_piece': type_piece,
                                'nom_type_piece': nom_type_piece,
                                'dn': dn_int,
                                'nom_dn': nom_dn,
                                'quantite': quantite_float,
                                'surface_unitaire': surface_unitaire,
                                'surface_totale': surface_totale,
                            }

                            elements_sablage.append(nouvel_element)
                            request.session['elements_sablage'] = elements_sablage

                            messages.success(request, f'{nom_type_piece} {nom_dn} ajouté avec succès!')
                        else:
                            messages.error(request,
                                           'Cette combinaison type/diamètre n\'est pas disponible pour le sablage.')
                    else:
                        messages.error(request, 'Combinaison type de pièce/diamètre invalide.')

                except (ValueError, TypeError):
                    messages.error(request, 'Veuillez entrer des valeurs valides.')

        elif 'supprimer_element' in request.POST:
            # Supprimer un élément
            try:
                index_to_remove = int(request.POST.get('element_index'))
                if 0 <= index_to_remove < len(elements_sablage):
                    removed_element = elements_sablage.pop(index_to_remove)
                    request.session['elements_sablage'] = elements_sablage
                    messages.success(request, f'{removed_element["nom_type_piece"]} supprimé.')
            except (ValueError, TypeError, IndexError):
                messages.error(request, 'Erreur lors de la suppression.')

        elif 'calculer_final' in request.POST:
            # Calculer le coût total et créer l'élément d'estimation
            if elements_sablage:
                # Calculer la surface globale
                surface_globale = sum(elem['surface_totale'] for elem in elements_sablage)
                prix_total_sablage = surface_globale * PRIX_SABLAGE_M2

                # Créer un résumé des éléments
                resume_elements = []
                for elem in elements_sablage:
                    resume_elements.append(f"{elem['nom_type_piece']} {elem['nom_dn']} (Qté: {elem['quantite']:g})")

                designation = f"Sablage tuyauterie - {', '.join(resume_elements)}"
                caracteristiques = f"Surface totale: {surface_globale:.3f} m² - Prix: {PRIX_SABLAGE_M2:,} CFA/m²"

                # Créer l'élément d'estimation ou demande personnalisée
                try:
                    # Option 1: Créer directement un EstimationElement
                    EstimationElement.objects.create(
                        projet=projet,
                        element=None,  # Pas d'élément standard
                        quantite=surface_globale,  # La quantité sera la surface en m²
                        prix_unitaire_fixe=PRIX_SABLAGE_M2,
                        # On pourrait ajouter des champs pour stocker les détails
                    )

                    # Option 2: Ou créer une demande personnalisée pré-approuvée
                    # discipline_sablage = Discipline.objects.filter(code__icontains='sablage').first()
                    # if not discipline_sablage:
                    #     discipline_sablage = Discipline.objects.first()  # Fallback

                    # DemandeElement.objects.create(
                    #     projet=projet,
                    #     categorie=categorie,
                    #     discipline=discipline_sablage,
                    #     designation=designation,
                    #     caracteristiques=caracteristiques,
                    #     unite='m2',
                    #     quantite=surface_globale,
                    #     statut='approuve',
                    #     prix_unitaire_admin=PRIX_SABLAGE_M2,
                    #     date_validation=timezone.now()
                    # )

                    # Nettoyer la session
                    del request.session['elements_sablage']

                    # Mettre à jour le résumé
                    summary, created = EstimationSummary.objects.get_or_create(projet=projet)
                    summary.calculer_totaux()

                    messages.success(request,
                                     f'Sablage tuyauterie ajouté: {surface_globale:.3f} m² - {prix_total_sablage:,.2f} CFA')
                    return redirect('category_selection')

                except Exception as e:
                    messages.error(request, f'Erreur lors de l\'enregistrement: {str(e)}')
            else:
                messages.warning(request, 'Aucun élément à calculer.')

    # Calculer le total en cours
    surface_globale_temp = sum(elem['surface_totale'] for elem in elements_sablage)
    prix_total_temp = surface_globale_temp * PRIX_SABLAGE_M2

    context.update({
        'surface_globale': surface_globale_temp,
        'prix_total': prix_total_temp,
        'nb_elements': len(elements_sablage),
    })

    return render(request, 'client/sablage_tuyauterie.html', context)


def ajax_calculer_surface_sablage(request):
    """Calcul AJAX de la surface pour aperçu en temps réel"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            type_piece = data.get('type_piece')
            dn = int(data.get('dn', 0))
            quantite = float(data.get('quantite', 0))

            # Mêmes données que dans la vue principale
            SURFACES_SABLAGE = {
                # ... (reprendre les mêmes données)
            }

            if dn in SURFACES_SABLAGE and type_piece in SURFACES_SABLAGE[dn]:
                surface_unitaire = SURFACES_SABLAGE[dn][type_piece]
                surface_totale = quantite * surface_unitaire

                return JsonResponse({
                    'success': True,
                    'surface_unitaire': surface_unitaire,
                    'surface_totale': surface_totale,
                    'disponible': surface_unitaire > 0
                })
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Combinaison non disponible'
                })

        except (ValueError, TypeError, KeyError):
            return JsonResponse({
                'success': False,
                'message': 'Données invalides'
            })

    return JsonResponse({'success': False})


def ajax_calculer_surface_sablage(request):
    """Calcul AJAX de la surface pour aperçu en temps réel"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            type_piece = data.get('type_piece')
            dn = int(data.get('dn', 0))
            quantite = float(data.get('quantite', 0))

            # Données complètes des surfaces unitaires (basées sur vos tableaux Excel)
            SURFACES_SABLAGE = {
                15: {  # DN 15 (1/2") - 21,3 mm
                    'tube': 0.067, 'coude_90': 0.004, 'coude_45': 0.002,
                    'coude_90_r5d': 0.007, 'coude_secteur': 0, 'te': 0,
                    'bride': 0, 'reduction': 0, 'cap': 0
                },
                20: {  # DN 20 (3/4") - 26,7 mm
                    'tube': 0.084, 'coude_90': 0.005, 'coude_45': 0.003,
                    'coude_90_r5d': 0.013, 'coude_secteur': 0, 'te': 0,
                    'bride': 0, 'reduction': 0, 'cap': 0
                },
                25: {  # DN 25 (1") - 33,4 mm
                    'tube': 0.105, 'coude_90': 0.006, 'coude_45': 0.003,
                    'coude_90_r5d': 0.021, 'coude_secteur': 0, 'te': 0,
                    'bride': 0, 'reduction': 0, 'cap': 0
                },
                32: {  # DN 32 (1" 1/4) - 42,2 mm
                    'tube': 0.133, 'coude_90': 0.010, 'coude_45': 0.005,
                    'coude_90_r5d': 0.033, 'coude_secteur': 0, 'te': 0,
                    'bride': 0, 'reduction': 0, 'cap': 0
                },
                40: {  # DN 40 (1" 1/2) - 48,3 mm
                    'tube': 0.152, 'coude_90': 0.014, 'coude_45': 0.007,
                    'coude_90_r5d': 0.045, 'coude_secteur': 0, 'te': 0,
                    'bride': 0, 'reduction': 0, 'cap': 0
                },
                50: {  # DN 50 (2") - 60,3 mm
                    'tube': 0.189, 'coude_90': 0.023, 'coude_45': 0.011,
                    'coude_90_r5d': 0.076, 'coude_secteur': 0, 'te': 0,
                    'bride': 0, 'reduction': 0, 'cap': 0
                },
                65: {  # DN 65 (2" 1/2) - 73 mm
                    'tube': 0.229, 'coude_90': 0.034, 'coude_45': 0.017,
                    'coude_90_r5d': 0.114, 'coude_secteur': 0, 'te': 0,
                    'bride': 0, 'reduction': 0, 'cap': 0
                },
                80: {  # DN 80 (3") - 88,9 mm
                    'tube': 0.279, 'coude_90': 0.050, 'coude_45': 0.025,
                    'coude_90_r5d': 0.167, 'coude_secteur': 0, 'te': 0,
                    'bride': 0, 'reduction': 0, 'cap': 0
                },
                100: {  # DN 100 (4") - 114,3 mm
                    'tube': 0.359, 'coude_90': 0.086, 'coude_45': 0.043,
                    'coude_90_r5d': 0.287, 'coude_secteur': 0, 'te': 0,
                    'bride': 0, 'reduction': 0, 'cap': 0
                },
                125: {  # DN 125 (5") - 139,7 mm
                    'tube': 0.439, 'coude_90': 0.131, 'coude_45': 0.066,
                    'coude_90_r5d': 0.438, 'coude_secteur': 0, 'te': 0,
                    'bride': 0, 'reduction': 0, 'cap': 0
                },
                150: {  # DN 150 (6") - 168,3 mm
                    'tube': 0.529, 'coude_90': 0.190, 'coude_45': 0.095,
                    'coude_90_r5d': 0.633, 'coude_secteur': 0.194, 'te': 0,
                    'bride': 0, 'reduction': 0, 'cap': 0
                },
                200: {  # DN 200 (8") - 219,1 mm
                    'tube': 0.688, 'coude_90': 0.330, 'coude_45': 0.165,
                    'coude_90_r5d': 1.099, 'coude_secteur': 0.338, 'te': 0,
                    'bride': 0, 'reduction': 0, 'cap': 0
                },
                250: {  # DN 250 (10") - 273 mm
                    'tube': 0.858, 'coude_90': 0.513, 'coude_45': 0.257,
                    'coude_90_r5d': 1.711, 'coude_secteur': 0.525, 'te': 0,
                    'bride': 0, 'reduction': 0, 'cap': 0
                },
                300: {  # DN 300 (12") - 323,9 mm
                    'tube': 1.018, 'coude_90': 0.731, 'coude_45': 0.365,
                    'coude_90_r5d': 2.436, 'coude_secteur': 0.748, 'te': 0,
                    'bride': 0, 'reduction': 0, 'cap': 0
                },
                350: {  # DN 350 (14") - 355,6 mm
                    'tube': 1.117, 'coude_90': 0.936, 'coude_45': 0.468,
                    'coude_90_r5d': 3.120, 'coude_secteur': 0.958, 'te': 0,
                    'bride': 0, 'reduction': 0, 'cap': 0
                },
                400: {  # DN 400 (16") - 406,5 mm
                    'tube': 1.277, 'coude_90': 1.222, 'coude_45': 0.611,
                    'coude_90_r5d': 4.076, 'coude_secteur': 1.252, 'te': 0,
                    'bride': 0, 'reduction': 0, 'cap': 0
                },
                450: {  # DN 450 (18") - 457,2 mm
                    'tube': 1.436, 'coude_90': 1.547, 'coude_45': 0.774,
                    'coude_90_r5d': 5.158, 'coude_secteur': 1.585, 'te': 0,
                    'bride': 0, 'reduction': 0, 'cap': 0
                },
                500: {  # DN 500 (20") - 508 mm
                    'tube': 1.596, 'coude_90': 1.910, 'coude_45': 0.955,
                    'coude_90_r5d': 6.368, 'coude_secteur': 1.953, 'te': 0,
                    'bride': 0, 'reduction': 0, 'cap': 0
                },
                600: {  # DN 600 (24") - 609,6 mm
                    'tube': 1.915, 'coude_90': 2.751, 'coude_45': 1.376,
                    'coude_90_r5d': 9.169, 'coude_secteur': 2.815, 'te': 0,
                    'bride': 0, 'reduction': 0, 'cap': 0
                },
                700: {  # DN 700 (28") - 711 mm
                    'tube': 2.234, 'coude_90': 3.744, 'coude_45': 1.872,
                    'coude_90_r5d': 12.48, 'coude_secteur': 3.283, 'te': 0,
                    'bride': 0, 'reduction': 0, 'cap': 0
                },
                750: {  # DN 750 (30") - 762 mm
                    'tube': 0, 'coude_90': 4.298, 'coude_45': 2.149,
                    'coude_90_r5d': 14.33, 'coude_secteur': 0, 'te': 0,
                    'bride': 0, 'reduction': 0, 'cap': 0
                },
                800: {  # DN 800 (32") - 813 mm
                    'tube': 2.554, 'coude_90': 4.892, 'coude_45': 2.446,
                    'coude_90_r5d': 16.30, 'coude_secteur': 5.003, 'te': 0,
                    'bride': 0, 'reduction': 0, 'cap': 0
                },
                900: {  # DN 900 (36") - 914 mm
                    'tube': 2.871, 'coude_90': 6.185, 'coude_45': 3.093,
                    'coude_90_r5d': 20.62, 'coude_secteur': 6.330, 'te': 0,
                    'bride': 0, 'reduction': 0, 'cap': 0
                },
                1000: {  # DN 1000 (40") - 1016 mm
                    'tube': 3.192, 'coude_90': 7.641, 'coude_45': 3.821,
                    'coude_90_r5d': 25.47, 'coude_secteur': 7.823, 'te': 0,
                    'bride': 0, 'reduction': 0, 'cap': 0
                }
            }

            if dn in SURFACES_SABLAGE and type_piece in SURFACES_SABLAGE[dn]:
                surface_unitaire = SURFACES_SABLAGE[dn][type_piece]
                surface_totale = quantite * surface_unitaire

                return JsonResponse({
                    'success': True,
                    'surface_unitaire': surface_unitaire,
                    'surface_totale': surface_totale,
                    'disponible': surface_unitaire > 0,
                    'dn_mm': {  # Correspondance DN vers diamètre externe en mm
                        15: 21.3, 20: 26.7, 25: 33.4, 32: 42.2, 40: 48.3,
                        50: 60.3, 65: 73, 80: 88.9, 100: 114.3, 125: 139.7,
                        150: 168.3, 200: 219.1, 250: 273, 300: 323.9, 350: 355.6,
                        400: 406.5, 450: 457.2, 500: 508, 600: 609.6, 700: 711,
                        750: 762, 800: 813, 900: 914, 1000: 1016
                    }.get(dn, dn)
                })
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Combinaison non disponible',
                    'disponible': False
                })

        except (ValueError, TypeError, KeyError) as e:
            return JsonResponse({
                'success': False,
                'message': f'Données invalides: {str(e)}',
                'disponible': False
            })

    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'})

################
# estimation/views.py - Vues d'authentification

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.hashers import make_password, check_password
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from .models import Client
import datetime


def client_login(request):
    """Connexion client"""
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')

        if not email or not password:
            messages.error(request, 'Veuillez remplir tous les champs.')
            return render(request, 'client/auth/login.html')

        try:
            client = Client.objects.get(email=email, actif=True)
            if client.check_password(password):
                # Connexion réussie
                request.session['client_id'] = client.id
                request.session['client_nom'] = client.nom
                client.last_login = datetime.datetime.now()
                client.save(update_fields=['last_login'])

                messages.success(request, f'Bienvenue {client.nom} !')

                # Rediriger vers la page demandée ou l'index
                next_url = request.GET.get('next', 'index')
                return redirect(next_url)
            else:
                messages.error(request, 'Email ou mot de passe incorrect.')
        except Client.DoesNotExist:
            messages.error(request, 'Email ou mot de passe incorrect.')

    return render(request, 'client/auth/login.html')


def client_register(request):
    """Inscription client"""
    if request.method == 'POST':
        nom = request.POST.get('nom', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        password_confirm = request.POST.get('password_confirm', '')
        telephone = request.POST.get('telephone', '').strip()
        entreprise = request.POST.get('entreprise', '').strip()
        adresse = request.POST.get('adresse', '').strip()

        # Validation
        errors = []

        if not nom:
            errors.append('Le nom est obligatoire.')

        if not email:
            errors.append('L\'email est obligatoire.')
        else:
            try:
                validate_email(email)
                # Vérifier l'unicité
                if Client.objects.filter(email=email).exists():
                    errors.append('Cette adresse email est déjà utilisée.')
            except ValidationError:
                errors.append('Format d\'email invalide.')

        if not password:
            errors.append('Le mot de passe est obligatoire.')
        elif len(password) < 6:
            errors.append('Le mot de passe doit contenir au moins 6 caractères.')

        if password != password_confirm:
            errors.append('Les mots de passe ne correspondent pas.')

        if errors:
            for error in errors:
                messages.error(request, error)
            return render(request, 'client/auth/register.html', {
                'form_data': request.POST
            })

        # Création du client
        try:
            client = Client.objects.create(
                nom=nom,
                email=email,
                telephone=telephone,
                entreprise=entreprise,
                adresse=adresse
            )
            client.set_password(password)
            client.save()

            messages.success(request, 'Votre compte a été créé avec succès ! Vous pouvez maintenant vous connecter.')
            return redirect('client_login')

        except Exception as e:
            messages.error(request, 'Une erreur est survenue lors de la création du compte.')

    return render(request, 'client/auth/register.html')


def client_logout(request):
    """Déconnexion client"""
    if 'client_id' in request.session:
        client_nom = request.session.get('client_nom', '')
        # Nettoyer la session
        request.session.flush()
        messages.success(request, f'À bientôt {client_nom} !')

    return redirect('client_login')


def client_profile(request):
    """Profil client"""
    client_id = request.session.get('client_id')
    if not client_id:
        return redirect('client_login')

    client = get_object_or_404(Client, id=client_id)

    if request.method == 'POST':
        nom = request.POST.get('nom', '').strip()
        telephone = request.POST.get('telephone', '').strip()
        entreprise = request.POST.get('entreprise', '').strip()
        adresse = request.POST.get('adresse', '').strip()

        if not nom:
            messages.error(request, 'Le nom est obligatoire.')
        else:
            client.nom = nom
            client.telephone = telephone
            client.entreprise = entreprise
            client.adresse = adresse
            client.save()

            # Mettre à jour la session
            request.session['client_nom'] = client.nom

            messages.success(request, 'Profil mis à jour avec succès !')

    return render(request, 'client/auth/profile.html', {'client': client})


def client_change_password(request):
    """Changement de mot de passe"""
    client_id = request.session.get('client_id')
    if not client_id:
        return redirect('client_login')

    client = get_object_or_404(Client, id=client_id)

    if request.method == 'POST':
        current_password = request.POST.get('current_password', '')
        new_password = request.POST.get('new_password', '')
        new_password_confirm = request.POST.get('new_password_confirm', '')

        # Validation
        if not client.check_password(current_password):
            messages.error(request, 'Mot de passe actuel incorrect.')
        elif len(new_password) < 6:
            messages.error(request, 'Le nouveau mot de passe doit contenir au moins 6 caractères.')
        elif new_password != new_password_confirm:
            messages.error(request, 'Les nouveaux mots de passe ne correspondent pas.')
        else:
            client.set_password(new_password)
            client.save()
            messages.success(request, 'Mot de passe modifié avec succès !')
            return redirect('client_profile')

    return render(request, 'client/auth/change_password.html')


# Décorateur pour protéger les vues
def client_required(view_func):
    """Décorateur pour vérifier l'authentification client"""

    def _wrapped_view(request, *args, **kwargs):
        if not request.session.get('client_id'):
            messages.warning(request, 'Veuillez vous connecter pour accéder à cette page.')
            return redirect('client_login')
        return view_func(request, *args, **kwargs)

    return _wrapped_view


# Modifier les vues existantes pour utiliser l'authentification

@client_required
def index(request):
    """Page d'accueil pour les clients authentifiés"""
    client_id = request.session.get('client_id')
    client = get_object_or_404(Client, id=client_id)

    # Statistiques personnalisées pour le client
    projets_client = Projet.objects.filter(client=client, actif=True)
    projets_actifs = projets_client.count()
    categories = Categorie.objects.all().count()
    elements = Element.objects.filter(actif=True).count()

    context = {
        'client': client,
        'projets_recents': projets_client.order_by('-date_creation')[:5],
        'stats': {
            'projets': projets_actifs,
            'categories': categories,
            'elements': elements,
        }
    }
    return render(request, 'client/index.html', context)


@client_required
def project_selection(request):
    """Sélection d'un projet existant ou création d'un nouveau"""
    client_id = request.session.get('client_id')
    client = get_object_or_404(Client, id=client_id)

    # Projets du client connecté uniquement
    projets = Projet.objects.filter(client=client, actif=True)

    if request.method == 'POST':
        if 'nouveau_projet' in request.POST:
            nom_projet = request.POST.get('nom_projet')
            description = request.POST.get('description', '')

            if nom_projet:
                projet = Projet.objects.create(
                    nom=nom_projet,
                    description=description,
                    client=client,
                    client_nom=client.nom  # Compatibilité
                )
                EstimationSummary.objects.create(projet=projet)
                request.session['projet_id'] = projet.id
                messages.success(request, f'Projet "{nom_projet}" créé avec succès!')
                return redirect('category_selection')
            else:
                messages.error(request, 'Le nom du projet est obligatoire.')

        elif 'projet_existant' in request.POST:
            projet_id = request.POST.get('projet_id')
            if projet_id:
                # Vérifier que le projet appartient au client
                try:
                    projet = Projet.objects.get(id=projet_id, client=client)
                    request.session['projet_id'] = int(projet_id)
                    return redirect('category_selection')
                except Projet.DoesNotExist:
                    messages.error(request, 'Projet non trouvé.')

    return render(request, 'client/project_selection.html', {
        'projets': projets,
        'client': client
    })


# Middleware pour ajouter le client au contexte
class ClientMiddleware:
    def __init__(self, get_response):
        self.get_response = get_response

    def __call__(self, request):
        # Ajouter le client au request si connecté
        client_id = request.session.get('client_id')
        if client_id:
            try:
                request.client = Client.objects.get(id=client_id, actif=True)
            except Client.DoesNotExist:
                # Client supprimé ou désactivé, nettoyer la session
                request.session.flush()
                request.client = None
        else:
            request.client = None

        response = self.get_response(request)
        return response

# imports nécessaires en haut du fichier
from django.views.decorators.http import require_POST
from django.db import transaction

@client_required
@require_POST
def supprimer_projet(request, projet_id):
    """Permet au client connecté de supprimer un de ses projets."""
    client_id = request.session.get('client_id')
    projet = get_object_or_404(Projet, id=projet_id, client_id=client_id)

    nom = projet.nom
    with transaction.atomic():
        # OPTION A — suppression définitive (cascade sur EstimationElement, DemandeElement, EstimationSummary, etc.)
        projet.delete()

        # OPTION B — "archivage" (soft delete). Si tu préfères, remplace les 2 lignes au-dessus par :
        # projet.actif = False
        # projet.save(update_fields=['actif'])

        # nettoyer la session si le projet supprimé était sélectionné
        if request.session.get('projet_id') == projet_id:
            request.session.pop('projet_id', None)
            request.session.pop('elements_sablage', None)

    messages.success(request, f'Le projet "{nom}" a été supprimé.')
    return redirect('project_selection')
